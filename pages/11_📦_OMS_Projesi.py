import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
import json
from io import BytesIO
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side

# Merkezi token sistemini import et
from token_manager import (
    check_token_charge,
    charge_token,
    render_token_widget,
    get_token_balance
)

st.set_page_config(page_title="Depo Birleştirme - Proje Yönetimi", layout="wide", page_icon="📦")

# ==============================================
# AUTHENTICATION KONTROLÜ
# ==============================================

if "authenticated" not in st.session_state or not st.session_state.authenticated:
    st.error("❌ Bu sayfaya erişmek için giriş yapmalısınız!")
    st.info("👉 Lütfen ana sayfadan giriş yapın.")
    
    col1, col2, col3 = st.columns([1, 1, 1])
    with col2:
        if st.button("🏠 Ana Sayfaya Dön", use_container_width=True, type="primary"):
            st.switch_page("Home.py")
    
    st.stop()

# ==============================================
# TOKEN KONTROLÜ
# ==============================================

username = st.session_state.user_info["username"]
module_name = "oms_proje"

should_charge = check_token_charge(username, module_name)

if should_charge:
    success, remaining, message = charge_token(username, module_name)
    
    if not success:
        st.error(f"❌ {message}")
        st.error("Token bakiyeniz tükendi!")
        st.stop()
    else:
        st.session_state.user_info["remaining_tokens"] = remaining
        
        if remaining <= 10:
            st.warning(f"⚠️ Token azalıyor! Kalan: {remaining}")

# ==============================================
# SIDEBAR
# ==============================================

with st.sidebar:
    # Kullanıcı profili
    st.markdown(f"""
    <div style='padding: 20px; background: linear-gradient(135deg, #667eea 0%, #764ba2 100%); 
                border-radius: 10px; margin-bottom: 20px; text-align: center;'>
        <div style='font-size: 3rem; margin-bottom: 10px;'>👤</div>
        <div style='color: white; font-size: 1.2rem; font-weight: 600;'>{st.session_state.user_info['name']}</div>
        <div style='color: rgba(255,255,255,0.8); font-size: 0.9rem;'>{st.session_state.user_info['title']}</div>
        <div style='margin-top: 10px; padding: 5px 10px; background: rgba(255,255,255,0.2); 
                    border-radius: 20px; display: inline-block; color: white; font-size: 0.85rem;'>
            {st.session_state.user_info['role'].upper()}
        </div>
    </div>
    """, unsafe_allow_html=True)
    
    # Token widget (merkezi)
    render_token_widget(username)
    
    st.markdown("---")
    
    # Navigasyon
    st.markdown("### 🧭 Navigasyon")
    if st.button("🏠 Ana Sayfa", use_container_width=True):
        st.switch_page("Home.py")
    
    st.markdown("---")
    
    # Çıkış
    if st.button("🚪 Çıkış Yap", use_container_width=True):
        st.session_state.authenticated = False
        st.session_state.user_info = None
        st.switch_page("Home.py")

# ==============================================
# ANA UYGULAMA BAŞLANGIÇ
# ==============================================

# Session state'de proje verilerini sakla
if 'proje_verileri' not in st.session_state:
    st.session_state.proje_verileri = {
        "FAZ 0: ANALİZ": {
            "baslangic": 0, "sure": 3, "renk": "🔴", "durum": "Planlandı",
            "gorevler": [
                {"id": "0.1", "gorev": "EVE Ürün Portföyü Analizi", "aciklama": "Hangi ürünler paketlenebilir?", 
                 "sure": 1, "baslangic_hafta": 1, "sorumlu": "Ertuğrul + Gökhan + Tayfun", "oncelik": "Yüksek", "durum": "Planlandı"},
                {"id": "0.2", "gorev": "Paket Tipi Belirleme", "aciklama": "Display box tipleri", 
                 "sure": 1, "baslangic_hafta": 2, "sorumlu": "Ertuğrul + Gökhan + Ali Akçay", "oncelik": "Yüksek", "durum": "Planlandı"},
                {"id": "0.3", "gorev": "Ürün Sınıflandırma", "aciklama": "Palet/Koli/Açık adet", 
                 "sure": 1, "baslangic_hafta": 2, "sorumlu": "Ertuğrul + Gökhan + Ferhat", "oncelik": "Kritik", "durum": "Planlandı"},
                {"id": "0.4", "gorev": "Maliyet Analizi", "aciklama": "Mevcut maliyetler", 
                 "sure": 1, "baslangic_hafta": 3, "sorumlu": "Finans + Ertuğrul + Gökhan", "oncelik": "Yüksek", "durum": "Planlandı"},
                {"id": "0.5", "gorev": "ROI Analizi", "aciklama": "Paketleme yatırımı", 
                 "sure": 1, "baslangic_hafta": 3, "sorumlu": "Finans + Ertuğrul + Gökhan", "oncelik": "Yüksek", "durum": "Planlandı"},
                {"id": "0.6", "gorev": "Veri Toplama", "aciklama": "6 ay geçmiş veri", 
                 "sure": 2, "baslangic_hafta": 1, "sorumlu": "Ertuğrul + Gökhan + Özcan", "oncelik": "Kritik", "durum": "Planlandı"},
                {"id": "0.7", "gorev": "Kapasite Analizi", "aciklama": "Ana depo kapasite", 
                 "sure": 1, "baslangic_hafta": 3, "sorumlu": "Ertuğrul + Ferhat", "oncelik": "Orta", "durum": "Planlandı"}
            ]
        },
        "FAZ 1: SİSTEM": {
            "baslangic": 3, "sure": 6, "renk": "🟢", "durum": "Planlandı",
            "gorevler": [
                {"id": "1.1", "gorev": "Simülasyon Modülü", "aciklama": "Ne olurdu analizi", 
                 "sure": 2, "baslangic_hafta": 4, "sorumlu": "Ertuğrul + Gökhan + Özcan", "oncelik": "Kritik", "durum": "Planlandı"},
                {"id": "1.2", "gorev": "Koli Bozma Algoritması", "aciklama": "Otomatik hesaplama", 
                 "sure": 2, "baslangic_hafta": 4, "sorumlu": "Ertuğrul + Gökhan + Özcan", "oncelik": "Kritik", "durum": "Planlandı"},
                {"id": "1.3", "gorev": "Transfer Sistemi", "aciklama": "Açık adet transfer", 
                 "sure": 2, "baslangic_hafta": 6, "sorumlu": "Ertuğrul + Gökhan + Özcan", "oncelik": "Yüksek", "durum": "Planlandı"},
                {"id": "1.4", "gorev": "Açık Adet Dashboard", "aciklama": "Görünürlük sistemi", 
                 "sure": 2, "baslangic_hafta": 6, "sorumlu": "Ertuğrul + Gökhan + Özcan", "oncelik": "Yüksek", "durum": "Planlandı"},
                {"id": "1.5", "gorev": "Önceliklendirme", "aciklama": "FIFO/FEFO", 
                 "sure": 2, "baslangic_hafta": 8, "sorumlu": "Ertuğrul + Gökhan + Özcan", "oncelik": "Orta", "durum": "Planlandı"},
                {"id": "1.6", "gorev": "Sevk Kural Motoru", "aciklama": "7 kural sistemi", 
                 "sure": 3, "baslangic_hafta": 6, "sorumlu": "Ertuğrul + Gökhan + Özcan", "oncelik": "Kritik", "durum": "Planlandı"},
                {"id": "1.7", "gorev": "Sevk Algoritması", "aciklama": "Otomatik öneri", 
                 "sure": 2, "baslangic_hafta": 8, "sorumlu": "Ertuğrul + Gökhan + Özcan", "oncelik": "Yüksek", "durum": "Planlandı"},
                {"id": "1.8", "gorev": "Sevk Dashboard", "aciklama": "Manuel onay", 
                 "sure": 2, "baslangic_hafta": 8, "sorumlu": "Ertuğrul + Gökhan + Özcan", "oncelik": "Yüksek", "durum": "Planlandı"},
                {"id": "1.9", "gorev": "Entegrasyon Test", "aciklama": "Tüm modüller", 
                 "sure": 1, "baslangic_hafta": 9, "sorumlu": "Ertuğrul + Gökhan + Özcan", "oncelik": "Kritik", "durum": "Planlandı"}
            ]
        },
        "FAZ 2: PİLOT": {
            "baslangic": 9, "sure": 15, "renk": "🔵", "durum": "Planlandı",
            "gorevler": [
                {"id": "2.1", "gorev": "Pilot Seçimi", "aciklama": "Kategoriler", 
                 "sure": 1, "baslangic_hafta": 10, "sorumlu": "Ertuğrul + Gökhan", "oncelik": "Kritik", "durum": "Planlandı"},
                {"id": "2.2", "gorev": "EVE Paketleme", "aciklama": "İlk parti", 
                 "sure": 2, "baslangic_hafta": 11, "sorumlu": "Tayfun + Ali Akçay", "oncelik": "Yüksek", "durum": "Planlandı"},
                {"id": "2.3", "gorev": "Stok Transferi", "aciklama": "Ana depoya", 
                 "sure": 1, "baslangic_hafta": 13, "sorumlu": "Ertuğrul + Ferhat", "oncelik": "Yüksek", "durum": "Planlandı"},
                {"id": "2.4", "gorev": "Pilot 1. Ay", "aciklama": "Canlı test", 
                 "sure": 4, "baslangic_hafta": 14, "sorumlu": "Ertuğrul + Gökhan", "oncelik": "Kritik", "durum": "Planlandı"},
                {"id": "2.5", "gorev": "Optimizasyon", "aciklama": "İyileştirmeler", 
                 "sure": 1, "baslangic_hafta": 18, "sorumlu": "Ertuğrul + Gökhan + Özcan", "oncelik": "Kritik", "durum": "Planlandı"},
                {"id": "2.6", "gorev": "Faz 2A Geçiş", "aciklama": "Orta hacim", 
                 "sure": 3, "baslangic_hafta": 19, "sorumlu": "Ertuğrul + Gökhan", "oncelik": "Yüksek", "durum": "Planlandı"},
                {"id": "2.7", "gorev": "EVE %50", "aciklama": "Yaygınlaştırma", 
                 "sure": 6, "baslangic_hafta": 19, "sorumlu": "Tayfun + Ali Akçay", "oncelik": "Yüksek", "durum": "Planlandı"},
                {"id": "2.8", "gorev": "Tam Geçiş", "aciklama": "Tüm kategoriler", 
                 "sure": 2, "baslangic_hafta": 22, "sorumlu": "Ertuğrul + Gökhan", "oncelik": "Kritik", "durum": "Planlandı"},
                {"id": "2.9", "gorev": "Depo Kararı", "aciklama": "Kapat/Küçült", 
                 "sure": 1, "baslangic_hafta": 24, "sorumlu": "Yönetim", "oncelik": "Kritik", "durum": "Planlandı"},
                {"id": "2.10", "gorev": "Depo Düzenleme", "aciklama": "Tasfiye", 
                 "sure": 1, "baslangic_hafta": 24, "sorumlu": "Ertuğrul", "oncelik": "Orta", "durum": "Planlandı"}
            ]
        },
        "FAZ 3: OMS MAĞAZA OPTİMİZASYONU": {
            "baslangic": 24, "sure": 12, "renk": "🟡", "durum": "Planlandı",
            "gorevler": [
                {"id": "3.1", "gorev": "Mevcut Mağaza Ağı Analizi", "aciklama": "Mağaza sayısı, dağılım, satış performansı analizi", 
                 "sure": 2, "baslangic_hafta": 25, "sorumlu": "Ertuğrul + Gökhan + Volkan", "oncelik": "Kritik", "durum": "Planlandı"},
                {"id": "3.2", "gorev": "Personel ve Kargo Maliyet Modelleme", "aciklama": "Mağaza başına personel maliyeti, kargo vs e-ticaret maliyet karşılaştırması", 
                 "sure": 2, "baslangic_hafta": 25, "sorumlu": "Ertuğrul + Gökhan + Finans", "oncelik": "Kritik", "durum": "Planlandı"},
                {"id": "3.3", "gorev": "Matematiksel Optimizasyon Modeli", "aciklama": "Mağaza karlılığı, başabaş noktası, ROI hesaplamaları", 
                 "sure": 3, "baslangic_hafta": 27, "sorumlu": "Ertuğrul + Gökhan", "oncelik": "Kritik", "durum": "Planlandı"},
                {"id": "3.4", "gorev": "Bölge Bazlı Talep Analizi", "aciklama": "Hangi bölgelerde mağaza eksik? Pazar potansiyeli nedir?", 
                 "sure": 2, "baslangic_hafta": 27, "sorumlu": "Ertuğrul + Gökhan + Pazarlama", "oncelik": "Yüksek", "durum": "Planlandı"},
                {"id": "3.5", "gorev": "İstisna Mağaza Belirleme", "aciklama": "Hangi mağazalar stratejik? (Franchise, flagship, yüksek trafik)", 
                 "sure": 1, "baslangic_hafta": 29, "sorumlu": "Volkan + Ertuğrul + Gökhan", "oncelik": "Yüksek", "durum": "Planlandı"},
                {"id": "3.6", "gorev": "Yeni Mağaza Açılış Senaryoları", "aciklama": "Kaç mağaza, hangi lokasyonlar, yatırım tutarları", 
                 "sure": 2, "baslangic_hafta": 30, "sorumlu": "Ertuğrul + Gökhan + Finans", "oncelik": "Yüksek", "durum": "Planlandı"},
                {"id": "3.7", "gorev": "ISO In-Store Ordering Ürün Segmentasyonu", "aciklama": "Hangi ürünler büyük/pahalı? ISO'ya yönlendirme kriterleri", 
                 "sure": 2, "baslangic_hafta": 30, "sorumlu": "Ertuğrul + Gökhan + Volkan", "oncelik": "Kritik", "durum": "Planlandı"},
                {"id": "3.8", "gorev": "ISO Teşvik Mekanizması Tasarımı", "aciklama": "Mağaza personeline indirim/komisyon, müşteriye özel fırsatlar", 
                 "sure": 2, "baslangic_hafta": 32, "sorumlu": "Volkan + Pazarlama + Finans", "oncelik": "Yüksek", "durum": "Planlandı"},
                {"id": "3.9", "gorev": "OMS Kural Motoru ISO Entegrasyonu", "aciklama": "Büyük ürünlerde otomatik ISO önerisi, stok yönlendirme", 
                 "sure": 2, "baslangic_hafta": 32, "sorumlu": "Ertuğrul + Gökhan + Özcan", "oncelik": "Kritik", "durum": "Planlandı"},
                {"id": "3.10", "gorev": "Mağazacılık ile El Sıkışma Toplantıları", "aciklama": "Veri paylaşımı, itirazlara matematiksel cevaplar, ortak karar", 
                 "sure": 2, "baslangic_hafta": 34, "sorumlu": "Yönetim + Ertuğrul + Gökhan + Volkan", "oncelik": "Kritik", "durum": "Planlandı"},
                {"id": "3.11", "gorev": "Mağaza Açılış Pilot Projesi", "aciklama": "Seçilen 2-3 lokasyonda pilot mağaza açılışı", 
                 "sure": 3, "baslangic_hafta": 34, "sorumlu": "Volkan", "oncelik": "Yüksek", "durum": "Planlandı"},
                {"id": "3.12", "gorev": "ISO Teşvik Pilotu", "aciklama": "Seçili mağazalarda ISO teşvik kampanyası, sonuç ölçümü", 
                 "sure": 2, "baslangic_hafta": 35, "sorumlu": "Volkan + Pazarlama", "oncelik": "Yüksek", "durum": "Planlandı"},
                {"id": "3.13", "gorev": "Sonuç Değerlendirme ve Strateji Finalizasyonu", "aciklama": "Pilot sonuçları, nihai mağaza sayısı kararı, ISO hedefleri", 
                 "sure": 1, "baslangic_hafta": 36, "sorumlu": "Yönetim + Tüm Ekip", "oncelik": "Kritik", "durum": "Planlandı"}
            ]
        }
    }

if 'baslangic_tarihi' not in st.session_state:
    st.session_state.baslangic_tarihi = datetime.now()

st.title("📦 DEPO BİRLEŞTİRME PROJESİ")
st.subheader("İnteraktif Proje Yönetim Sistemi")
st.markdown("---")

fazlar = st.session_state.proje_verileri
toplam_gorev = sum(len(faz['gorevler']) for faz in fazlar.values())
toplam_sure = max(faz['baslangic'] + faz['sure'] for faz in fazlar.values())

col1, col2, col3, col4 = st.columns(4)
col1.metric("Toplam Görev", toplam_gorev)
col2.metric("Proje Süresi", f"{toplam_sure} Hafta")
col3.metric("Faz Sayısı", len(fazlar))

# Durum özeti
tamamlanan = sum(1 for faz in fazlar.values() for g in faz['gorevler'] if g['durum'] == 'Tamamlandı')
devam_eden = sum(1 for faz in fazlar.values() for g in faz['gorevler'] if g['durum'] == 'Devam Ediyor')
col4.metric("Tamamlanan", f"{tamamlanan}/{toplam_gorev}")

st.markdown("---")

# Tarih seçici
col1, col2 = st.columns([3, 1])
with col1:
    yeni_tarih = st.date_input(
        "📅 Proje Başlangıç Tarihi",
        value=st.session_state.baslangic_tarihi
    )
    if yeni_tarih != st.session_state.baslangic_tarihi.date():
        st.session_state.baslangic_tarihi = datetime.combine(yeni_tarih, datetime.min.time())
        st.rerun()

baslangic = st.session_state.baslangic_tarihi
bitis = baslangic + timedelta(weeks=toplam_sure)

with col2:
    st.info(f"**Bitiş:** {bitis.strftime('%d.%m.%Y')}")

# Ana Tabs
tab1, tab2, tab3, tab4, tab5 = st.tabs([
    "📋 Yönetim", 
    "📊 Gantt Chart", 
    "✏️ Düzenle", 
    "➕ Ekle",
    "📥 Veri"
])

# TAB 1: YÖNETİM
with tab1:
    st.header("🎯 Proje Yönetim Ekranı")
    
    for faz_adi, faz in fazlar.items():
        with st.expander(f"{faz['renk']} **{faz_adi}** - {faz['durum']}", expanded=True):
            col1, col2, col3, col4 = st.columns(4)
            col1.markdown(f"**Başlangıç:** H{faz['baslangic']+1}")
            col2.markdown(f"**Süre:** {faz['sure']} hafta")
            col3.markdown(f"**Görev:** {len(faz['gorevler'])}")
            with col4:
                yeni_durum = st.selectbox(
                    "Faz Durumu",
                    ["Planlandı", "Devam Ediyor", "Tamamlandı", "Beklemede"],
                    index=["Planlandı", "Devam Ediyor", "Tamamlandı", "Beklemede"].index(faz['durum']),
                    key=f"faz_{faz_adi}"
                )
                if yeni_durum != faz['durum']:
                    st.session_state.proje_verileri[faz_adi]['durum'] = yeni_durum
                    st.rerun()
            
            st.progress(faz['sure'] / toplam_sure)
            
            data = []
            for g in faz['gorevler']:
                bas = baslangic + timedelta(weeks=g['baslangic_hafta']-1)
                bit = bas + timedelta(weeks=g['sure'])
                data.append({
                    'ID': g['id'],
                    'Görev': g['gorev'],
                    'Açıklama': g['aciklama'],
                    'Süre': g['sure'],
                    'Başlangıç H': g['baslangic_hafta'],
                    'Başlangıç': bas.strftime('%d.%m'),
                    'Bitiş': bit.strftime('%d.%m'),
                    'Sorumlu': g['sorumlu'],
                    'Öncelik': g['oncelik'],
                    'Durum': g['durum']
                })
            
            df = pd.DataFrame(data)
            st.dataframe(df, use_container_width=True, hide_index=True)
            st.markdown("---")

# TAB 2: GANTT CHART
with tab2:
    st.header("📊 Gantt Chart - Proje Timeline")
    
    # Gantt Chart CSS stilleri
    st.markdown("""
    <style>
    .gantt-container {
        overflow-x: auto;
        margin: 20px 0;
    }
    .gantt-table {
        border-collapse: collapse;
        width: 100%;
        font-size: 12px;
    }
    .gantt-table th {
        background-color: #f0f2f6;
        padding: 8px;
        text-align: left;
        border: 1px solid #ddd;
        position: sticky;
        top: 0;
        z-index: 10;
    }
    .gantt-table td {
        padding: 6px;
        border: 1px solid #ddd;
        text-align: center;
    }
    .gantt-task {
        height: 25px;
        border-radius: 4px;
        display: inline-block;
        position: relative;
        margin: 2px 0;
    }
    .tamamlandi { background-color: #4caf50; color: white; }
    .devam-ediyor { background-color: #ff9800; color: white; }
    .planlandi { background-color: #2196f3; color: white; }
    .beklemede { background-color: #9e9e9e; color: white; }
    .gorev-info {
        font-size: 10px;
        padding: 2px 8px;
        white-space: nowrap;
        overflow: hidden;
        text-overflow: ellipsis;
    }
    .faz-header {
        background-color: #e3f2fd;
        font-weight: bold;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # Gantt Chart HTML oluştur
    max_hafta = 40
    
    gantt_html = '<div class="gantt-container"><table class="gantt-table">'
    
    # Header - Haftalar
    gantt_html += '<tr><th style="min-width:200px;">Görev</th><th style="min-width:80px;">ID</th><th style="min-width:100px;">Sorumlu</th><th style="min-width:80px;">Durum</th>'
    for h in range(1, max_hafta + 1):
        gantt_html += f'<th style="min-width:40px;">H{h}</th>'
    gantt_html += '</tr>'
    
    # Her faz ve görev için satır
    for faz_adi, faz in fazlar.items():
        # Faz başlığı
        gantt_html += f'<tr class="faz-header"><td colspan="{max_hafta + 4}">{faz["renk"]} {faz_adi}</td></tr>'
        
        # Görevler
        for gorev in faz['gorevler']:
            durum_class = gorev['durum'].lower().replace(' ', '-').replace('ı', 'i')
            
            gantt_html += f'<tr>'
            gantt_html += f'<td style="text-align:left;">{gorev["gorev"][:40]}</td>'
            gantt_html += f'<td>{gorev["id"]}</td>'
            gantt_html += f'<td>{gorev["sorumlu"]}</td>'
            gantt_html += f'<td><span class="gantt-task {durum_class}" style="width:60px; display:inline-block;">{gorev["durum"][:4]}</span></td>'
            
            # Hafta hücreleri
            bas_h = gorev['baslangic_hafta']
            sure = gorev['sure']
            
            for h in range(1, max_hafta + 1):
                if h >= bas_h and h < bas_h + sure:
                    if h == bas_h:
                        colspan = min(sure, max_hafta - h + 1)
                        gantt_html += f'<td colspan="{colspan}">'
                        gantt_html += f'<div class="gantt-task {durum_class}" style="width:100%;">'
                        gantt_html += f'<span class="gorev-info">{gorev["id"]}</span>'
                        gantt_html += '</div></td>'
                        for _ in range(1, colspan):
                            continue
                elif h < bas_h or h >= bas_h + sure:
                    if h not in range(bas_h, bas_h + sure):
                        gantt_html += '<td></td>'
            
            gantt_html += '</tr>'
    
    gantt_html += '</table></div>'
    
    st.markdown(gantt_html, unsafe_allow_html=True)
    
    st.markdown("---")
    # Bu kodu TAB 2: GANTT CHART'ın sonuna ekle
# st.markdown("### 📌 Durum Göstergeleri") satırından ÖNCE

    # ============================================
    # EXCEL EXPORT BÖLÜMÜ
    # ============================================
    
    st.markdown("---")
    st.subheader("📥 Gantt Chart İndir")
    
    col1, col2, col3 = st.columns(3)
    
    with col1:
        # Excel export - Gantt formatında
        if st.button("📊 Excel İndir (Gantt Formatı)", use_container_width=True, type="primary"):
            # Excel workbook oluştur
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Gantt Chart"
            
            # Başlık satırı
            headers = ['Görev', 'ID', 'Sorumlu', 'Durum'] + [f'H{h}' for h in range(1, max_hafta + 1)]
            ws.append(headers)
            
            # Başlık stili
            header_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
            header_font = Font(bold=True, size=11)
            
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Renkler - Durum bazında
            durum_colors = {
                'Tamamlandı': '4CAF50',      # Yeşil
                'Devam Ediyor': 'FF9800',    # Turuncu
                'Planlandı': '2196F3',       # Mavi
                'Beklemede': '9E9E9E'        # Gri
            }
            
            # Her faz için satırlar ekle
            row_num = 2
            for faz_adi, faz in fazlar.items():
                # Faz başlığı
                faz_row = [f"{faz['renk']} {faz_adi}"] + [''] * (max_hafta + 3)
                ws.append(faz_row)
                
                # Faz satırını birleştir ve stil ver
                ws.merge_cells(f'A{row_num}:{openpyxl.utils.get_column_letter(max_hafta + 4)}{row_num}')
                faz_cell = ws[f'A{row_num}']
                faz_cell.fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
                faz_cell.font = Font(bold=True, size=12)
                faz_cell.alignment = Alignment(horizontal='left', vertical='center')
                row_num += 1
                
                # Görevler
                for gorev in faz['gorevler']:
                    gorev_row = [
                        gorev['gorev'],
                        gorev['id'],
                        gorev['sorumlu'],
                        gorev['durum']
                    ]
                    
                    # Hafta hücreleri
                    bas_h = gorev['baslangic_hafta']
                    sure = gorev['sure']
                    
                    for h in range(1, max_hafta + 1):
                        if h >= bas_h and h < bas_h + sure:
                            gorev_row.append(gorev['id'])  # Görev ID'sini yaz
                        else:
                            gorev_row.append('')
                    
                    ws.append(gorev_row)
                    
                    # Görev satırına stil ver
                    current_row = ws[row_num]
                    
                    # Durum hücresine renk
                    durum_cell = current_row[3]
                    durum_color = durum_colors.get(gorev['durum'], 'FFFFFF')
                    durum_cell.fill = PatternFill(start_color=durum_color, end_color=durum_color, fill_type="solid")
                    durum_cell.font = Font(color="FFFFFF", bold=True)
                    durum_cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    # Gantt hücrelerine renk
                    for col_idx in range(5, max_hafta + 5):
                        cell = current_row[col_idx - 1]
                        if cell.value == gorev['id']:
                            cell.fill = PatternFill(start_color=durum_color, end_color=durum_color, fill_type="solid")
                            cell.font = Font(color="FFFFFF", size=8, bold=True)
                            cell.alignment = Alignment(horizontal='center', vertical='center')
                    
                    row_num += 1
            
            # Kolon genişlikleri
            ws.column_dimensions['A'].width = 45
            ws.column_dimensions['B'].width = 8
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 15
            for col_idx in range(5, max_hafta + 5):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 4
            
            # Border ekle - Tüm hücrelere
            thin_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )
            
            for row in ws.iter_rows(min_row=1, max_row=row_num-1, min_col=1, max_col=max_hafta+4):
                for cell in row:
                    cell.border = thin_border
            
            # Satır yüksekliği
            for row in range(1, row_num):
                ws.row_dimensions[row].height = 25
            
            # Excel'i kaydet
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            st.download_button(
                label="📥 Gantt_Chart.xlsx İndir",
                data=output,
                file_name=f"Gantt_Chart_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
    
    with col2:
        # CSV export - Basit tablo
        if st.button("📋 CSV İndir (Tablo)", use_container_width=True):
            tum_gorevler = []
            for faz_adi, faz in fazlar.items():
                for g in faz['gorevler']:
                    bas = baslangic + timedelta(weeks=g['baslangic_hafta']-1)
                    bit = bas + timedelta(weeks=g['sure'])
                    tum_gorevler.append({
                        'Faz': faz_adi,
                        'ID': g['id'],
                        'Görev': g['gorev'],
                        'Açıklama': g['aciklama'],
                        'Süre (Hafta)': g['sure'],
                        'Başlangıç Haftası': g['baslangic_hafta'],
                        'Başlangıç Tarihi': bas.strftime('%d.%m.%Y'),
                        'Bitiş Tarihi': bit.strftime('%d.%m.%Y'),
                        'Sorumlu': g['sorumlu'],
                        'Öncelik': g['oncelik'],
                        'Durum': g['durum']
                    })
            
            df_export = pd.DataFrame(tum_gorevler)
            csv = df_export.to_csv(index=False, encoding='utf-8-sig')
            
            st.download_button(
                label="📥 Gorev_Listesi.csv İndir",
                data=csv,
                file_name=f"Gorev_Listesi_{datetime.now().strftime('%Y%m%d')}.csv",
                mime="text/csv",
                use_container_width=True
            )
    
    with col3:
        # Özet rapor - 3 sayfa Excel
        if st.button("📊 Özet Rapor (Excel)", use_container_width=True):
            output = BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Sayfa 1: Faz özeti
                faz_data = []
                for faz_adi, faz in fazlar.items():
                    tamamlanan_gorev = sum(1 for g in faz['gorevler'] if g['durum'] == 'Tamamlandı')
                    toplam_gorev = len(faz['gorevler'])
                    faz_data.append({
                        'Faz': faz_adi,
                        'Başlangıç Haftası': faz['baslangic'],
                        'Süre (Hafta)': faz['sure'],
                        'Toplam Görev': toplam_gorev,
                        'Tamamlanan': tamamlanan_gorev,
                        'Tamamlanma %': round(tamamlanan_gorev / toplam_gorev * 100, 1) if toplam_gorev > 0 else 0,
                        'Durum': faz['durum']
                    })
                
                df_faz = pd.DataFrame(faz_data)
                df_faz.to_excel(writer, sheet_name='Faz Özeti', index=False)
                
                # Sayfa 2: Tüm görevler
                tum_gorevler = []
                for faz_adi, faz in fazlar.items():
                    for g in faz['gorevler']:
                        bas = baslangic + timedelta(weeks=g['baslangic_hafta']-1)
                        bit = bas + timedelta(weeks=g['sure'])
                        tum_gorevler.append({
                            'Faz': faz_adi,
                            'ID': g['id'],
                            'Görev': g['gorev'],
                            'Açıklama': g['aciklama'],
                            'Süre': g['sure'],
                            'Başlangıç H': g['baslangic_hafta'],
                            'Başlangıç': bas.strftime('%d.%m.%Y'),
                            'Bitiş': bit.strftime('%d.%m.%Y'),
                            'Sorumlu': g['sorumlu'],
                            'Öncelik': g['oncelik'],
                            'Durum': g['durum']
                        })
                
                df_gorev = pd.DataFrame(tum_gorevler)
                df_gorev.to_excel(writer, sheet_name='Tüm Görevler', index=False)
                
                # Sayfa 3: Durum özeti
                durum_data = []
                for durum in ['Planlandı', 'Devam Ediyor', 'Tamamlandı', 'Beklemede']:
                    count = sum(1 for faz in fazlar.values() for g in faz['gorevler'] if g['durum'] == durum)
                    durum_data.append({'Durum': durum, 'Görev Sayısı': count})
                
                df_durum = pd.DataFrame(durum_data)
                df_durum.to_excel(writer, sheet_name='Durum Özeti', index=False)
            
            output.seek(0)
            
            st.download_button(
                label="📥 Proje_Ozet_Raporu.xlsx İndir",
                data=output,
                file_name=f"Proje_Ozet_Raporu_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
    st.markdown("### 📌 Durum Göstergeleri")
    col1, col2, col3, col4 = st.columns(4)
    col1.markdown("🟢 **Tamamlandı** - Yeşil")
    col2.markdown("🟠 **Devam Ediyor** - Turuncu")
    col3.markdown("🔵 **Planlandı** - Mavi")
    col4.markdown("⚫ **Beklemede** - Gri")

# TAB 3: DÜZENLE
with tab3:
    st.header("✏️ Görev Düzenleme")
    
    faz_sec = st.selectbox("Faz Seçin", list(fazlar.keys()), key="edit_faz_select")
    
    if faz_sec:
        gorev_options = {}
        for g in fazlar[faz_sec]['gorevler']:
            gorev_options[g['id']] = f"{g['id']} - {g['gorev']}"
        
        if gorev_options:
            gorev_sec_id = st.selectbox(
                "Düzenlenecek Görevi Seçin",
                options=list(gorev_options.keys()),
                format_func=lambda x: gorev_options[x],
                key="edit_gorev_select"
            )
            
            if gorev_sec_id:
                gorev = next((g for g in fazlar[faz_sec]['gorevler'] if g['id'] == gorev_sec_id), None)
                
                if gorev:
                    st.markdown("---")
                    st.subheader(f"Görev: {gorev['id']} - {gorev['gorev']}")
                    
                    with st.form(f"edit_form_{gorev['id'].replace('.', '_')}"):
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            yeni_gorev = st.text_input("Görev Adı", value=gorev['gorev'])
                            yeni_aciklama = st.text_area("Açıklama", value=gorev['aciklama'])
                            yeni_sure = st.number_input("Süre (hafta)", min_value=1, value=gorev['sure'])
                            yeni_bas_h = st.number_input("Başlangıç Haftası", min_value=1, value=gorev['baslangic_hafta'])
                        
                        with col2:
                            yeni_sorumlu = st.text_input("Sorumlu", value=gorev['sorumlu'])
                            yeni_oncelik = st.selectbox(
                                "Öncelik",
                                ["Kritik", "Yüksek", "Orta", "Düşük"],
                                index=["Kritik", "Yüksek", "Orta", "Düşük"].index(gorev['oncelik'])
                            )
                            yeni_durum = st.selectbox(
                                "Durum",
                                ["Planlandı", "Devam Ediyor", "Tamamlandı", "Beklemede"],
                                index=["Planlandı", "Devam Ediyor", "Tamamlandı", "Beklemede"].index(gorev['durum'])
                            )
                            yeni_id = st.text_input("ID (dikkatli değiştirin)", value=gorev['id'])
                        
                        col1, col2 = st.columns(2)
                        kaydet = col1.form_submit_button("💾 Kaydet", use_container_width=True)
                        sil = col2.form_submit_button("🗑️ Sil", use_container_width=True)
                        
                        if kaydet:
                            idx = next((i for i, g in enumerate(fazlar[faz_sec]['gorevler']) if g['id'] == gorev_sec_id), None)
                            
                            if idx is not None:
                                st.session_state.proje_verileri[faz_sec]['gorevler'][idx] = {
                                    'id': yeni_id,
                                    'gorev': yeni_gorev,
                                    'aciklama': yeni_aciklama,
                                    'sure': yeni_sure,
                                    'baslangic_hafta': yeni_bas_h,
                                    'sorumlu': yeni_sorumlu,
                                    'oncelik': yeni_oncelik,
                                    'durum': yeni_durum
                                }
                                st.success("✅ Görev kaydedildi!")
                                st.rerun()
                        
                        if sil:
                            st.session_state.proje_verileri[faz_sec]['gorevler'] = [
                                g for g in fazlar[faz_sec]['gorevler'] if g['id'] != gorev_sec_id
                            ]
                            st.success(f"✅ Görev {gorev_sec_id} silindi!")
                            st.rerun()
        else:
            st.info("Bu fazda henüz görev yok.")

# TAB 4: EKLE
with tab4:
    st.header("➕ Yeni Ekle")
    
    tip = st.radio("Ne eklemek istiyorsunuz?", ["Görev", "Faz"])
    
    if tip == "Görev":
        st.subheader("Yeni Görev Ekle")
        
        with st.form("yeni_gorev"):
            hedef = st.selectbox("Hangi Faza Eklenecek?", list(fazlar.keys()))
            
            col1, col2 = st.columns(2)
            with col1:
                yeni_id = st.text_input("Görev ID", placeholder="Örn: 0.8 veya 1.10")
                yeni_gorev = st.text_input("Görev Adı", placeholder="Görev başlığı")
                yeni_aciklama = st.text_area("Açıklama", placeholder="Detaylı açıklama")
                yeni_sure = st.number_input("Süre (hafta)", min_value=1, value=1)
            
            with col2:
                yeni_bas_h = st.number_input("Başlangıç Haftası", min_value=1, value=1)
                yeni_sorumlu = st.text_input("Sorumlu", placeholder="Örn: Ertuğrul + Gökhan")
                yeni_oncelik = st.selectbox("Öncelik", ["Kritik", "Yüksek", "Orta", "Düşük"])
                yeni_durum = st.selectbox("Durum", ["Planlandı", "Devam Ediyor", "Tamamlandı", "Beklemede"])
            
            if st.form_submit_button("➕ Görevi Ekle", use_container_width=True):
                if yeni_id and yeni_gorev:
                    mevcut_idler = [g['id'] for g in fazlar[hedef]['gorevler']]
                    if yeni_id in mevcut_idler:
                        st.error(f"⚠️ {yeni_id} ID'si zaten kullanılıyor!")
                    else:
                        st.session_state.proje_verileri[hedef]['gorevler'].append({
                            'id': yeni_id,
                            'gorev': yeni_gorev,
                            'aciklama': yeni_aciklama,
                            'sure': yeni_sure,
                            'baslangic_hafta': yeni_bas_h,
                            'sorumlu': yeni_sorumlu,
                            'oncelik': yeni_oncelik,
                            'durum': yeni_durum
                        })
                        st.success(f"✅ Yeni görev '{yeni_gorev}' ({yeni_id}) eklendi!")
                        st.rerun()
                else:
                    st.error("⚠️ Görev ID ve Görev Adı zorunludur!")
    
    else:
        st.subheader("Yeni Faz Ekle")
        
        with st.form("yeni_faz"):
            faz_adi = st.text_input("Faz Adı", placeholder="Örn: FAZ 4: YAYINLAMA")
            
            col1, col2, col3 = st.columns(3)
            with col1:
                faz_bas = st.number_input("Başlangıç Haftası", min_value=0, value=toplam_sure)
            with col2:
                faz_sure = st.number_input("Süre (hafta)", min_value=1, value=4)
            with col3:
                faz_renk = st.selectbox("Emoji", ["🔴", "🟢", "🔵", "🟡", "🟣", "⚫", "⚪", "🟤"])
            
            if st.form_submit_button("➕ Faz Ekle", use_container_width=True):
                if faz_adi:
                    if faz_adi in fazlar:
                        st.error(f"⚠️ '{faz_adi}' adında bir faz zaten var!")
                    else:
                        st.session_state.proje_verileri[faz_adi] = {
                            'baslangic': faz_bas,
                            'sure': faz_sure,
                            'renk': faz_renk,
                            'durum': 'Planlandı',
                            'gorevler': []
                        }
                        st.success(f"✅ Yeni faz '{faz_adi}' eklendi!")
                        st.rerun()
                else:
                    st.error("⚠️ Faz adı zorunludur!")

# TAB 5: VERİ İŞLEMLERİ
with tab5:
    st.header("📥 Veri İşlemleri")
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("📤 Dışa Aktar")
        
        json_data = json.dumps(st.session_state.proje_verileri, ensure_ascii=False, indent=2)
        st.download_button(
            "📥 JSON olarak İndir",
            json_data,
            "proje_verileri.json",
            "application/json",
            use_container_width=True
        )
        
        tum_gorevler = []
        for faz_adi, faz in fazlar.items():
            for g in faz['gorevler']:
                tum_gorevler.append({'Faz': faz_adi, **g})
        
        df_export = pd.DataFrame(tum_gorevler)
        st.download_button(
            "📥 CSV olarak İndir",
            df_export.to_csv(index=False).encode('utf-8'),
            "proje_gorevleri.csv",
            "text/csv",
            use_container_width=True
        )
    
    with col2:
        st.subheader("📥 İçe Aktar")
        
        uploaded = st.file_uploader("JSON Dosyası Yükle", type=['json'])
        if uploaded:
            try:
                data = json.loads(uploaded.read())
                st.success("✅ Dosya okundu!")
                
                if st.button("✅ Veriyi Projeye Yükle", use_container_width=True):
                    st.session_state.proje_verileri = data
                    st.success("✅ Proje verileri güncellendi!")
                    st.rerun()
            except Exception as e:
                st.error(f"❌ Hata: {str(e)}")
    
    st.markdown("---")
    st.subheader("🔄 Sıfırlama")
    
    if st.button("⚠️ Tüm Verileri Sıfırla ve Varsayılanlara Dön", type="secondary", use_container_width=True):
        if 'proje_verileri' in st.session_state:
            del st.session_state.proje_verileri
        st.success("✅ Proje verileri varsayılanlara döndürüldü!")
        st.rerun()

st.markdown("---")
st.caption(f"📦 Depo Birleştirme Projesi | Thorius AR4U | {datetime.now().strftime('%d.%m.%Y %H:%M')}")
